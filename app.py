from flask import Flask, render_template, request, jsonify, send_file
import numpy as np
import pickle
import json
import openpyxl
import io


with open('values.pkl', 'rb') as f:
    values = pickle.load(f)

with open('colors.pkl', 'rb') as f:
    colors = pickle.load(f)

with open('villages_list.pkl', 'rb') as f:
    villages_list = pickle.load(f)

with open('features_list.pkl', 'rb') as f:
    features_list = pickle.load(f)

polygons = []
with open("ru.json") as file:
    ru = json.load(file)
    for feature in ru['features']:
        if feature['properties']['id'] in ["RUBA"]:
            polygon = np.array(feature['geometry']['coordinates'][0])
            polygon = polygon[:, ::-1]
            polygons.append(polygon.tolist())

app = Flask(__name__)

@app.route("/")
def index():
    return render_template(
        "index.html",
        title = "Атлас русских говоров Башкирии",
        features_list=features_list,
        enumerate=enumerate
    )

@app.route('/get_villages', methods=['GET'])
def get_villages():
    return jsonify(villages_list)

@app.route('/get_features', methods=['GET'])
def get_features():
    return jsonify(features_list)

@app.route('/get_colors', methods=['GET'])
def get_colors():
    return jsonify(colors)

@app.route('/get_polygons', methods=['GET'])
def get_polygons():
    return jsonify(polygons)

@app.route('/get_values', methods=['POST'])
def get_values():
    features_id = request.get_json()["features_id"]
    values_subset = values[:, features_id]
    unique, inverse = np.unique(values_subset, axis=0, return_inverse=True)
    return jsonify(unique.tolist(), inverse.tolist())

@app.route('/download_features', methods=['POST'])
def download_features():
    village_id = request.get_json()["village_id"]
    features_subset = values[village_id].tolist()

    wb = openpyxl.Workbook()
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        wb.remove(sheet)

    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )

    for domain in features_list:
        dims = {}
        ws = wb.create_sheet(domain["domain"].replace(":", "."))
        for h, header in enumerate(["НОМЕР", "КАРТА", "ПРИЗНАКИ"], 1):
            cell = ws.cell(row=1, column=h)
            gray = "C0C0C0"
            cell.fill = openpyxl.styles.PatternFill(start_color=gray, end_color=gray, fill_type="solid")
            cell.font = openpyxl.styles.Font(bold=True)
            cell.border = thin_border
            cell.alignment = openpyxl.styles.Alignment(horizontal="left")

            cell.value = header

        for i, map_ in enumerate(domain["maps"], 2):
            presented_features = []
            for map_feature in map_["features"]:
                if features_subset[map_feature[0]]:
                    presented_features.append(map_feature[1])

            for v, val in enumerate([map_["number"], map_["name"], presented_features], 1):
                cell = ws.cell(row=i, column=v)
                cell.border = thin_border
                cell.alignment = openpyxl.styles.Alignment(horizontal="left")

                if v != 3:
                    cell.value = val
                else:
                    if val:
                        cell.value = "; ".join(val)

                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))

        dims[1] = 10
        for col, value in dims.items():
            ws.column_dimensions[["A", "B", "C"][col - 1]].width = value

    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name=f'{villages_list[village_id]["name"]}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/info')
def info():
    return render_template("info.html")

if __name__ == "__main__":
    app.run()